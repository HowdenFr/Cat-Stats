"""
This program is made to analyze an excel sheet from a penalty corner game delivered by 
the field hockey program. 
This is to help anyone who does a game easily analyze and send stats to the coaches. 
This program will have comments during it to remind the user as well as a reader understand
what is going on. 

Author: Frank Howden 

Big help to ChatGPT 



"""

import openpyxl 
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import os

#opening the field hockey excel sheets as well as the total season one. 

def select_excel_file():
    """
    Finds the excel sheet path through the windows folder. 
    This is only the path
    parameters: none
    returns: nothing
    
    
    """
    root=tk.Tk()
    root.withdraw() #take away a tk window. 

    file_path=filedialog.askopenfilename(
        title="Select a file",
        filetypes=[("Excel files","*.xlsx"),("Text files","*.txt"),("All files","*.*")],
        initialdir=".")
    
    if file_path: 
        
        root.destroy()
        return file_path
    else:
        
        root.destroy()
        return None
    


def excel_file_checker(file): 
    """
    checks if the excel sheet are good
    parameters: file. the file path to the excel sheet
    returns: workbook if you picked an excel sheet. 
            none: if it wasn't an excel sheet you picked. 

    """
    workbook=""
    try:
        if file.endswith('.xlsx') or file.endswith('.xls'):
            workbook=openpyxl.load_workbook(file)
            return workbook
        else:
            raise ValueError("unsupported file type.")
            
    except Exception as e:
        messagebox.showerror("error", f"{str(e)}")
    return None




def player_dictionary(player_text):
    """
    makes the player dictionary that will be used to store the stats for the players. 
    parameters: players: the .txt file that has all the players.
    returns: a dictionary of the players for offense and defense. 
    
    """
    offense_dictionary={}
    defense_dictionary={}
    try:
        with open(player_text,'r') as file:
            for line in file:
    
                name=line.strip()
                
                if name:
                    offense_dictionary[name]=[0]*10 #make up offense dictionary
            
                    defense_dictionary[name]=[0]*7 #make up defense dictionary
    except Exception as e:
        print(f"{e}")
    
    
    return offense_dictionary,defense_dictionary

def insert_stat_in_dictionary(name,dictionary_stats, index):
    """
    This function updates a stat at a certain position for each player. 
    This function is repeated multiple times, saves space. 
    Parameters: name: the player who got the stat
                dictionary_stats: the dictionary being update
                index: the index in the array
    
    """
    array=dictionary_stats[name]
    total=array[index]
    total=total+1
    array[index]=total
    dictionary_stats[name]=array
        
def offense_stat(game_file,offense_stats,offense_penalty_corner):
    """
    This function get's all the data that is needed to analyze into the offense dictionary
    Parameters: game_file: the file needed to scrape the stats
                offense_stats: the dictionary that keeps track of the players stats. 
                offense_penalty_corner: the stats for how many corners, recorners, and goals during a game. 
    
    """
    #this is getting the sheet for attack penalty

    attack_sheet=game_file['Attack']

    #getting who drawed
    for row in attack_sheet.iter_rows(min_row=2,max_row=attack_sheet.max_row,min_col=0, max_col=attack_sheet.max_column,values_only=True):
        offense_penalty_corner[0]+=1 #updating the amount of penalty corners
        
        for i in range(5,13): #updating the penalty corner lineup. 
            
            if(row[i] in offense_stats.keys()):
                index=i-4
                insert_stat_in_dictionary(row[i],offense_stats,index)
        if(row[3] in offense_stats.keys()): #finding who drew the penalty corner
            insert_stat_in_dictionary(row[3],offense_stats,0)
        if(row[16]=='Recorner'): #updating how many were penalty corners. 
            offense_penalty_corner[1]+=1
        if(row[19]=='Y'):
           
            offense_penalty_corner[2]+=1
    
   
  
    return offense_stats,offense_penalty_corner

def defense_stat(game_file, defense_stats,defense_penalty_corner):
    """
    This function get's all the data that is needed to analyze into the offense dictionary. 
    Parameters: game_file: the file needed to scrape the stats
                defense_stats: the dictionary that keeps track of defense penalty corners for the players
                defense_penalty_corner: the array that keeps track of the total corners, recorners, and goals in a game. 
    Returns: defense_stats, defense_penalty_corner

    """
     #this is getting the sheet for attack penalty

    defense_sheet=game_file['Defense']

    #getting who drawed
    for row in defense_sheet.iter_rows(min_row=2,max_row=defense_sheet.max_row,min_col=0, max_col=defense_sheet.max_column,values_only=True):
        defense_penalty_corner[0]+=1 #updating the amount of penalty corners
       
        for i in range(5,10): #updating the penalty corner lineup. 
            if(row[i] in defense_stats.keys()):
                index=i-4
                insert_stat_in_dictionary(row[i],defense_stats,index)
        
        if(row[3] in defense_stats.keys()): #finding who drew the penalty corner
            insert_stat_in_dictionary(row[3],defense_stats,0)
        
        if(row[18] in defense_stats.keys()): #inserting stat for who tipped a pass
            insert_stat_in_dictionary(row[18],defense_stats,6)
       
        if(row[19]=='Recorner'): #updating how many were penalty corners. 
            defense_penalty_corner[1]+=1
       
        if(row[20]=='Y'): #updating how many goals were let in
            defense_penalty_corner[2]+=1
    
    return defense_stats,defense_penalty_corner



def write_stats(game_file,OPC, DPC, game_file_path):
    """
    This function writes the stats for the in game sheet for the coaches
    Parameters: game_file: the file that has the sheets. 
    OPC: Offense_Penalty_corner array
                DPC: Defense_Penalty_Corner array 
                game_file_path: the game file path that will use to save the workbook. 
    Returns: Nothing
    """

    stat_sheet=game_file.create_sheet(title="Stats")
    #start inserting stats by row. 

    stat_sheet['A1']="Corner Execution"
    stat_sheet['A3']="Offense Recorner Percentage"
    
    if(OPC[0]==0): #can't divide by )
        
        stat_sheet['B1']=0
        stat_sheet['B3']=0
    else:
        stat_sheet['B1']=(OPC[2]/OPC[0]) * 100  #the percentage of offense execution
        stat_sheet['B3']=(OPC[1]/OPC[0]) * 100  #percentage of recorner on Offense Penalty Corners

    stat_sheet['A2']="Defense Corner Execution"
    stat_sheet['A4']="Defense Recorner Percentage"
    if(DPC[0]==0): #can't divide by zero
        stat_sheet['B2']=0
        stat_sheet['B4']=0
    else:
        
        stat_sheet['B2']=(DPC[2]/DPC[0]) * 100 #the percentage of defense execution
        stat_sheet['B4']=(DPC[1]/DPC[0]) * 100 #percentage of recorner on Defense Penalty Corner
    
    #all the numbers that go into the respected stats 
    stat_sheet['C1']=OPC[2]  
    stat_sheet['D1']=OPC[0]
    stat_sheet['C2']=DPC[2]
    stat_sheet['D2']=DPC[0]
    stat_sheet['C3']=OPC[1]
    stat_sheet['D3']=OPC[0]
    stat_sheet['C4']=DPC[1]
    stat_sheet['D4']=DPC[0]

    name_of_file=game_file_path.split('/')   
    game_file.save(name_of_file[len(name_of_file)-1])
    

def season_corner_stats(corner_sheet,OPC,DPC, season_file_path, season_file):
    """
    This function updates and writes the stats for the season file. It is a continious count
    Parameters: corner_sheet: the sheet that you will be updating
                OPC: offense penalty corner for a game
                DPC: defense penalty corners for a game
                season_file_path: the path to that season file. 
    Returns: Nothing
    """
    total_game_corners=OPC[0]
    total_defense_corners=DPC[0]
    #saving them as place holders 

    #computing the corner execution percentage on the season
    season_goals_score=int(corner_sheet['C1'].value)
    season_penalty_corners=int(corner_sheet['D1'].value)
    season_goals_score=season_goals_score + OPC[2]
    #updating the total penalty corners on the season
    season_penalty_corners+=total_game_corners
    corner_sheet['C1'].value=season_goals_score
    corner_sheet['D1'].value=season_penalty_corners
    corner_sheet['B1'].value=(season_goals_score/season_penalty_corners) * 100 

    #computing the defense penalty corner execution on the season
    season_goals_givenup=int(corner_sheet['C2'].value)
    season_penalty_corners_D=int(corner_sheet['D2'].value)
    season_goals_givenup+= DPC[2]
    #updting the total penalty corners on the season for defense
    season_penalty_corners_D+=total_defense_corners
    corner_sheet['C2'].value=season_goals_givenup
    corner_sheet['D2'].value=season_penalty_corners_D
    #calculating the season percentage
    corner_sheet['B2'].value=(season_goals_givenup/season_penalty_corners_D) * 100 

    #computing the offense recorners execution on the season
    offense_recorners=int(corner_sheet['C3'].value) + OPC[1]
    corner_sheet['C3'].value=offense_recorners
    corner_sheet['D3'].value=int(corner_sheet['D1'].value)
    #updating the recorner percentage on offense
    corner_sheet['B3'].value=offense_recorners/int(corner_sheet['D3'].value)
    
    #computing the offense recorners execution on the season
    defense_recorners=int(corner_sheet['C4'].value) + DPC[1]
    corner_sheet['C4'].value=defense_recorners
    corner_sheet['D4'].value=int(corner_sheet['D2'].value)
    #updating the recorner percentage on offense
    corner_sheet['B4'].value=defense_recorners/int(corner_sheet['D4'].value)


    season_file.save(season_file_path)


def attack_update(season_file,attack_sheet,off_stats,season_file_path):
    """
    This function updates and writes the stats for the attack penalty corners
    over the season. It is a continious count.
    Parameters: season_file: the workbook we are saving. 
                attack_sheet: the sheet we are updating
                off_stats: the offense stats for the game
                season_file_path: the file_path that will be used to save the file.
    Returns: Nothing
    """
    workbook=load_workbook(filename=season_file_path)
    attack_sheet=workbook['Attack']
    #go through all the rows

    for row in range(2, attack_sheet.max_row+1):
        
        #if the name in the row is in the dictionary, start updating
        
            name=str(attack_sheet.cell(row=row,column=1).value)
            
             #get the name and then the array from the dictionary
            
            if(name in off_stats.keys()):
                count=0
                array=off_stats[name]
        
                #find the cell and then update it with the dictionary values. 
                for col in range(2, attack_sheet.max_column + 1):
                    cell=attack_sheet.cell(row=row, column=col)
                    original_value=int(cell.value)
                    new_value=original_value + array[count]

                    cell.value=new_value
                    count+=1
    #save the file
                season_file.save(season_file_path)
                workbook.save(season_file_path)

def defense_update(season_file,defense_sheet,def_stats,season_file_path):
    """
    This function updates and writes the stats for the attack penalty corners
    over the season. It is a continious count.
    Parameters: season_file: the workbook we are saving
                defense_sheet: the sheet we are updating
                def_stats: the defense stats for the game
                season_file_path: the file_path that will be used to save the file.
    Returns: Nothing
    """
    workbook=load_workbook(filename=season_file_path)
    defense_sheet=workbook['Defense']
    for row in range(2,defense_sheet.max_row + 1):
        #if the name in the row is in the dictionary, start updating
        name=str(defense_sheet.cell(row=row,column=1).value)

        if (name in def_stats.keys()):
            
            #get the name and then the array from the dictionary
            count=0
            array=def_stats[name]
            
            #find the cell and then update it with the dictionary values. 
            for col in range(2,defense_sheet.max_column + 1 ):
                cell=defense_sheet.cell(row=row, column=col)
                
                original_value=int(cell.value)
                
                new_value=(original_value) + array[count]

                cell.value=new_value
                count+=1
                #save the file
            season_file.save(season_file_path)
            workbook.save(season_file_path)



           
            
            

    
    
def season_stats(season_file,off_stats,def_stats, OPC, DPC, season_file_path):
    """
    This function updates and writes the stats for the season file. It is a continious count
    Parameters: season_file: the file being used
                off_stats: offense stats in a dictionary for that game
                def_stats: defense stats in a dictionary for that game
                season_file_path: the path to that season file. 
    Returns: Nothing
    """
    name_of_sheet=season_file_path.split('/')
    name_of_sheet=name_of_sheet[len(name_of_sheet )-1]
    #open up the sheet we need. 
    
    season_corner_sheet=season_file['Season Stats']
    season_corner_stats(season_corner_sheet,OPC, DPC,name_of_sheet,season_file)

    #updating season stats

    attack=season_file['Attack']
    attack_update(season_file,attack,off_stats,name_of_sheet)

    defense=season_file['Defense']
    defense_update(season_file,defense,def_stats,name_of_sheet)


    


if __name__=="__main__":
    game_file_path=select_excel_file()
    
    season_file_path=select_excel_file()
    
    #checking if the files we selected are excel. 
    game_file=excel_file_checker(game_file_path)
    
    season_file=excel_file_checker(season_file_path)

    #finding the players sheet. It's a .txt file. 
    player_text=select_excel_file()
    #making the player dictionars
    offense_stats,defense_stats=player_dictionary(player_text)
    #analyzing the offense side of the excel sheet of the game file
    offense_penalty_corner=[0]*3

    defense_penalty_corner=[0]*3
    #the *3 is for goals scored, recorners, and total penalty corners

    offense_stats,offense_penalty_corner=offense_stat(game_file,offense_stats,offense_penalty_corner)
    defense_stats,defense_penalty_corner=defense_stat(game_file,defense_stats,defense_penalty_corner)
    write_stats(game_file,offense_penalty_corner,defense_penalty_corner, game_file_path)
    season_stats(season_file,offense_stats,defense_stats,offense_penalty_corner,defense_penalty_corner,season_file_path)







    

    





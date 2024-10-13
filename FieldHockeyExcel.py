"""
this program makes the excel sheet after each game

Author: Frank Howden 
"""

# through out the prgram these hashtags will help you understand what's going on. 
# full disclaimer ChatGPT is a massive help. 



from asyncio.windows_events import NULL
import pygsheets
import gspread
import openpyxl 
import aspose.words

class CreateStat():

    def __init__(self,offense_titles,offense_stats,defense_titles,defense_stats, other_stats,offense_total,defense_total):
        self.clean_up(other_stats,offense_stats,defense_stats)
        self.wb=openpyxl.Workbook()
        self.ws=self.wb.active
        self.ws.title="Davidson Field Hockey Match " + str(other_stats[0])
        self.name= other_stats[1]+" vs "+other_stats[2] + ".xlsx"
        
        
       
        self.insert_offense(offense_titles,offense_stats,offense_total)
        self.insert_defense(defense_titles,defense_stats,defense_total)
        self.wb.save(self.name)



    

        return
    
    def clean_up(self,other_stats,offense_stats,defense_stats):
        """
        this method cleans up the other stats
        Parameters: other_stats: the other_stats
        Returns: other_stats. 
    
        """
        date=other_stats[0]
        if('/' in date):
            date=date.replace('/','-')
        other_stats[0]=date
        
        if(defense_stats[0]==None):
            
            for i in range(len(defense_stats)):
                for m in range(len(defense_stats[i])):
                    defense_stats[i][m]="NA"
        return other_stats


   



    def insert_offense(self,offense_titles, offense_stats,offense_total):
        """
        this method inserts the offense sheet 
        Parameters: self
                    offense_titles: the titles for the offense stats
                    offense_stats: the stats for the offense
                    offense_total: the amount of penalty corners on offense that were played
        Returns: none

        
        """
        #this is the new worksheet we are working on

        self.ws=self.wb.create_sheet(title="Attack")
        
        current_row=1
        current_col=1
        #inserting the titles on the attack sheet
        for i in range(len(offense_titles)):
            self.ws.cell(row=current_row,column=current_col,value=offense_titles[i][0])

            current_col=current_col+1
        self.wb.save(self.name)
        
        current_col=1
        counter=0
    
        for v in range(offense_total):
            counter=0
            current_row=current_row+1
            current_col=1
            for row in offense_stats:
                
                counter=counter+1
               
                
                self.ws.cell(row=current_row,column=current_col,value=row[v])
                current_col=current_col+1
            self.wb.save(self.name)
        return
    
    def insert_defense(self,defense_titles,defense_stats,defense_total):
        """
        this method inserts the defense sheet 
        Parameters: self
                    defense_titles: the titles for the defense stats
                    defense_stats: the stats for the defense
                    defense_total: the amount of penalty corners on defense that were played
        Returns: none

        
        """
        #this is the new worksheet we are working on
        self.ws=self.wb.create_sheet(title="Defense")

        current_row=1
        current_col=1
        #inserting the titles on the defense sheet
        for m in range(len(defense_titles)):
            self.ws.cell(row=current_row, column=current_col, value=defense_titles[m][0])
            current_col=current_col +1
        
        current_col=1
        counter=0
        self.wb.save(self.name)
        for j in range(defense_total):
            current_row=current_row+1
            current_col=1
            counter=0
            for stat in defense_stats:
                
                counter=counter+1
                
                
                self.ws.cell(row=current_row,column=current_col,value=stat[j])
                current_col=current_col+1
            self.wb.save(self.name)
        return


        


